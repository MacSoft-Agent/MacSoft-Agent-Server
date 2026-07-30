from __future__ import annotations

import base64
import csv
from io import BytesIO, StringIO
from typing import Any

from macsoft.config import AppConfig
from macsoft.files.storage import AdminUploadedFileRecord, UploadedFileRecord, stored_path


MAX_EXTRACTED_TEXT_CHARS = 80_000
MAX_SPREADSHEET_ROWS = 2_000
MAX_SPREADSHEET_COLUMNS = 100


class AttachmentContentError(ValueError):
    def __init__(self, code: str, message: str) -> None:
        super().__init__(message)
        self.code = code


def _decode_text(data: bytes) -> str:
    for encoding in ("utf-8-sig", "utf-16", "cp1252"):
        try:
            return data.decode(encoding)
        except UnicodeDecodeError:
            continue
    raise AttachmentContentError(
        "file_text_unreadable",
        "The text file encoding could not be read safely.",
    )


def _extract_csv(data: bytes) -> str:
    rows: list[str] = []
    reader = csv.reader(StringIO(_decode_text(data)))
    for index, row in enumerate(reader):
        if index >= MAX_SPREADSHEET_ROWS:
            rows.append("[Additional rows omitted]")
            break
        rows.append("\t".join(str(value) for value in row[:MAX_SPREADSHEET_COLUMNS]))
    return "\n".join(rows)


def _extract_xlsx(data: bytes) -> str:
    try:
        from openpyxl import load_workbook
    except ImportError as error:
        raise AttachmentContentError(
            "spreadsheet_reader_unavailable",
            "The Server spreadsheet reader is not installed.",
        ) from error

    try:
        workbook = load_workbook(BytesIO(data), read_only=True, data_only=True)
    except Exception as error:
        raise AttachmentContentError(
            "spreadsheet_unreadable",
            "The XLSX workbook could not be read.",
        ) from error

    parts: list[str] = []
    remaining_rows = MAX_SPREADSHEET_ROWS
    try:
        for sheet in workbook.worksheets:
            if remaining_rows <= 0:
                parts.append("[Additional workbook rows omitted]")
                break
            parts.append(f"[Sheet: {sheet.title}]")
            for row in sheet.iter_rows(values_only=True):
                if remaining_rows <= 0:
                    parts.append("[Additional workbook rows omitted]")
                    break
                values = ["" if value is None else str(value) for value in row[:MAX_SPREADSHEET_COLUMNS]]
                parts.append("\t".join(values).rstrip())
                remaining_rows -= 1
    finally:
        workbook.close()
    return "\n".join(parts)


def _extract_pdf(data: bytes) -> str:
    try:
        from pypdf import PdfReader
    except ImportError as error:
        raise AttachmentContentError(
            "pdf_reader_unavailable",
            "The Server PDF reader is not installed.",
        ) from error

    try:
        reader = PdfReader(BytesIO(data))
        text = "\n\n".join((page.extract_text() or "").strip() for page in reader.pages).strip()
    except Exception as error:
        raise AttachmentContentError("pdf_unreadable", "The PDF could not be read.") from error
    if not text:
        raise AttachmentContentError(
            "scanned_pdf_requires_images",
            "This PDF has no readable text. Upload the relevant pages as JPG or PNG for visual extraction.",
        )
    return text


def _document_text(record: UploadedFileRecord | AdminUploadedFileRecord, data: bytes) -> str:
    if record.media_type == "text/csv":
        text = _extract_csv(data)
    elif record.media_type == "text/plain":
        text = _decode_text(data)
    elif record.media_type == "application/pdf":
        text = _extract_pdf(data)
    elif record.media_type == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet":
        text = _extract_xlsx(data)
    else:
        raise AttachmentContentError("unsupported_document", "This document type cannot be converted to text.")
    if len(text) > MAX_EXTRACTED_TEXT_CHARS:
        text = text[:MAX_EXTRACTED_TEXT_CHARS] + "\n[Additional document text omitted]"
    return text


def build_hermes_user_content(
    config: AppConfig,
    *,
    message: str,
    files: list[UploadedFileRecord | AdminUploadedFileRecord],
) -> str | list[dict[str, Any]]:
    if not files:
        return message

    text_parts = [message]
    image_parts: list[dict[str, Any]] = []
    for record in files:
        data = stored_path(config, record).read_bytes()
        if record.media_type.startswith("image/"):
            encoded = base64.b64encode(data).decode("ascii")
            image_parts.append(
                {
                    "type": "image_url",
                    "image_url": {
                        "url": f"data:{record.media_type};base64,{encoded}",
                        "detail": "high",
                    },
                }
            )
            text_parts.append(f"[Attached image: {record.original_name}]")
        else:
            extracted = _document_text(record, data)
            text_parts.append(
                "\n".join(
                    (
                        f"[BEGIN UNTRUSTED ATTACHMENT DATA: {record.original_name}]",
                        extracted,
                        f"[END UNTRUSTED ATTACHMENT DATA: {record.original_name}]",
                    )
                )
            )

    combined_text = "\n\n".join(part for part in text_parts if part)
    if not image_parts:
        return combined_text
    return [{"type": "text", "text": combined_text}, *image_parts]

